#!/usr/bin/env python3
"""
ЭТАП 1: Автоподсчёт T/F и MC
Вход:  Key.xlsx, Answers.xlsx
Выход: stage1_result.json (баллы Q1-20 + сырые ответы Q21-50)

Запуск:
  python3 stage1.py
"""

import openpyxl
import json
import sys

# ═══════════════════════════════════════════════════════════════
# ШАГ 1: Читаем ключ
# ═══════════════════════════════════════════════════════════════

def load_key(filepath):
    """
    Читает Key.xlsx и возвращает словарь вопросов.
    Автоматически определяет тип вопроса по формату ответа:
      - T или F → тип "tf"
      - A, B, C или D → тип "mc"
      - всё остальное → тип "open"
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]

    questions = {}

    for row in range(2, ws.max_row + 1):
        number = ws.cell(row=row, column=1).value
        if number is None:
            continue

        number = int(number)
        question_text = str(ws.cell(row=row, column=2).value or "")
        answer = str(ws.cell(row=row, column=3).value or "").strip()
        points = int(ws.cell(row=row, column=4).value or 0)
        comment = str(ws.cell(row=row, column=5).value or "").strip()

        # Определяем тип по формату ответа
        if answer.upper() in ("T", "F"):
            q_type = "tf"
        elif answer.upper() in ("A", "B", "C", "D"):
            q_type = "mc"
        else:
            q_type = "open"

        questions[number] = {
            "text": question_text,
            "answer": answer,
            "points": points,
            "type": q_type,
            "comment": comment
        }

    return questions


# ═══════════════════════════════════════════════════════════════
# ШАГ 2: Читаем ответы участников
# ═══════════════════════════════════════════════════════════════

def load_answers(filepath):
    """
    Читает Answers.xlsx и возвращает список участников.
    Ожидаемый формат: Name | Surname | Language | Q1 | Q2 | ... | Q50
    Номера вопросов извлекаются из заголовков (Q1 → 1, Q2 → 2, ...).
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]

    # Читаем заголовки и определяем, какой столбец = какой вопрос
    headers = []
    for col in range(1, ws.max_column + 1):
        headers.append(str(ws.cell(row=1, column=col).value or ""))

    # Маппинг: номер вопроса → номер столбца
    question_columns = {}
    for col_idx, header in enumerate(headers):
        if header.startswith("Q") and header[1:].isdigit():
            q_num = int(header[1:])
            question_columns[q_num] = col_idx + 1  # openpyxl считает с 1

    # Читаем участников
    participants = []

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if name is None:
            continue

        name = str(name).strip()
        surname = str(ws.cell(row=row, column=2).value or "").strip()
        language = str(ws.cell(row=row, column=3).value or "").strip()

        # Собираем все ответы
        answers = {}
        for q_num, col in question_columns.items():
            raw = ws.cell(row=row, column=col).value
            answer = str(raw).strip() if raw is not None else ""
            if answer in ("None", "-", "—", "–"):
                answer = ""
            answers[q_num] = answer

        participants.append({
            "name": name,
            "surname": surname,
            "language": language,
            "answers": answers
        })

    return participants


# ═══════════════════════════════════════════════════════════════
# ШАГ 3: Подсчёт баллов за T/F и MC
# ═══════════════════════════════════════════════════════════════

def score_auto_questions(participants, questions):
    """
    Для каждого участника считает баллы за вопросы типа tf и mc.
    TF: сравниваем T/F напрямую.
    MC: берём первую букву ответа участника, сравниваем с буквой из ключа.
    """
    for p in participants:
        p["scores"] = {}
        p["tf_total"] = 0
        p["mc_total"] = 0

        for q_num, q_info in questions.items():
            if q_info["type"] not in ("tf", "mc"):
                continue

            participant_answer = p["answers"].get(q_num, "").strip().upper()
            correct_answer = q_info["answer"].strip().upper()
            max_points = q_info["points"]

            if q_info["type"] == "tf":
                # Прямое сравнение: T с T, F с F
                is_correct = (participant_answer == correct_answer)

            elif q_info["type"] == "mc":
                # Берём первую букву ответа участника
                first_letter = participant_answer[0] if participant_answer else ""
                is_correct = (first_letter == correct_answer)

            earned = max_points if is_correct else 0

            p["scores"][q_num] = {
                "earned": earned,
                "max": max_points,
                "correct": is_correct,
                "participant_answer": p["answers"].get(q_num, ""),
                "correct_answer": q_info["answer"]
            }

            if q_info["type"] == "tf":
                p["tf_total"] += earned
            else:
                p["mc_total"] += earned

        p["auto_total"] = p["tf_total"] + p["mc_total"]

    return participants


# ═══════════════════════════════════════════════════════════════
# ШАГ 4: Вывод и сохранение
# ═══════════════════════════════════════════════════════════════

def print_results(participants, questions):
    """Печатает таблицу результатов Q1-20."""
    # Максимумы
    max_tf = sum(q["points"] for q in questions.values() if q["type"] == "tf")
    max_mc = sum(q["points"] for q in questions.values() if q["type"] == "mc")
    max_auto = max_tf + max_mc

    print(f"\n{'='*70}")
    print(f"ЭТАП 1: Автоподсчёт T/F и MC")
    print(f"{'='*70}")
    print(f"{'#':<4} {'Имя':<30} {'T/F':>6} {'MC':>6} {'Итого':>7}")
    print(f"{'-'*70}")

    # Сортируем по итогу
    sorted_p = sorted(participants, key=lambda x: x["auto_total"], reverse=True)

    for i, p in enumerate(sorted_p):
        full_name = f"{p['name']} {p['surname']}"
        print(f"{i+1:<4} {full_name:<30} {p['tf_total']:>3}/{max_tf:<2} "
              f"{p['mc_total']:>3}/{max_mc:<2} {p['auto_total']:>4}/{max_auto}")

    # Статистика
    avg = sum(p["auto_total"] for p in participants) / len(participants)
    print(f"{'-'*70}")
    print(f"Участников: {len(participants)} | Средний балл: {avg:.1f}/{max_auto}")


def save_result(participants, questions, filepath):
    """
    Сохраняет JSON с результатами.
    Включает: баллы Q1-20 + сырые ответы Q21-50 + метаданные ключа.
    """
    # Собираем открытые вопросы из ключа
    open_questions = {}
    for q_num, q_info in questions.items():
        if q_info["type"] == "open":
            open_questions[q_num] = q_info

    # Формируем данные для сохранения
    output = {
        "summary": {
            "total_participants": len(participants),
            "tf_questions": sum(1 for q in questions.values() if q["type"] == "tf"),
            "mc_questions": sum(1 for q in questions.values() if q["type"] == "mc"),
            "open_questions": sum(1 for q in questions.values() if q["type"] == "open"),
            "max_tf": sum(q["points"] for q in questions.values() if q["type"] == "tf"),
            "max_mc": sum(q["points"] for q in questions.values() if q["type"] == "mc"),
            "max_open": sum(q["points"] for q in questions.values() if q["type"] == "open"),
        },
        "open_key": {str(k): v for k, v in open_questions.items()},
        "participants": []
    }

    for p in participants:
        # Собираем только открытые ответы
        open_answers = {}
        for q_num in open_questions:
            open_answers[str(q_num)] = p["answers"].get(q_num, "")

        output["participants"].append({
            "name": p["name"],
            "surname": p["surname"],
            "full_name": f"{p['name']} {p['surname']}",
            "language": p["language"],
            "tf_total": p["tf_total"],
            "mc_total": p["mc_total"],
            "auto_total": p["auto_total"],
            "auto_scores": {str(k): v for k, v in p["scores"].items()},
            "open_answers": open_answers
        })

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\nСохранено: {filepath}")


# ═══════════════════════════════════════════════════════════════
# ЗАПУСК
# ═══════════════════════════════════════════════════════════════

def main():
    key_file = "Key.xlsx"
    answers_file = "Answers.xlsx"
    output_file = "stage1_result.json"

    # Шаг 1
    print("Читаю ключ...")
    questions = load_key(key_file)
    n_tf = sum(1 for q in questions.values() if q["type"] == "tf")
    n_mc = sum(1 for q in questions.values() if q["type"] == "mc")
    n_open = sum(1 for q in questions.values() if q["type"] == "open")
    print(f"  Вопросов: {len(questions)} ({n_tf} T/F, {n_mc} MC, {n_open} open)")

    # Шаг 2
    print("Читаю ответы...")
    participants = load_answers(answers_file)
    print(f"  Участников: {len(participants)}")

    # Шаг 3
    print("Считаю баллы Q1-20...")
    participants = score_auto_questions(participants, questions)

    # Шаг 4
    print_results(participants, questions)
    save_result(participants, questions, output_file)


if __name__ == "__main__":
    main()
